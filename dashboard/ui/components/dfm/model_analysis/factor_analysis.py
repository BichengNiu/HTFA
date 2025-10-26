# -*- coding: utf-8 -*-
"""
DFM因子分析组件

提供因子分析、可视化和结果展示功能
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import logging
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, Optional, Tuple, List
from scipy.cluster import hierarchy as sch

from dashboard.ui.components.dfm.base import DFMComponent, DFMServiceManager
from dashboard.core import get_global_dfm_manager
from dashboard.models.DFM.config import VisualizationDefaults


logger = logging.getLogger(__name__)


class FactorAnalysisComponent(DFMComponent):
    """DFM因子分析组件"""
    
    def __init__(self, service_manager: Optional[DFMServiceManager] = None):
        """
        初始化因子分析组件
        
        Args:
            service_manager: DFM服务管理器
        """
        super().__init__(service_manager)
        self._default_chart_settings = {
            'heatmap_colorscale': 'RdBu',
            'time_series_height': 400,
            'heatmap_height': 600
        }
    
    def get_component_id(self) -> str:
        """获取组件ID"""
        return "factor_analysis"
    
    def get_state_keys(self) -> list:
        """
        获取组件相关的状态键
        
        Returns:
            List[str]: 状态键列表
        """
        return [
            'dfm_analysis_results',
            'dfm_selected_factors',
            'dfm_chart_settings',
            'dfm_analysis_status'
        ]
    
    def validate_input(self, data: Dict) -> bool:
        """
        验证输入数据
        
        Args:
            data: 输入数据字典，包含模型数据和元数据
            
        Returns:
            bool: 验证是否通过
        """
        try:
            # 检查模型数据
            model_data = data.get('model_data')
            if model_data is None:
                logger.warning("缺少模型数据")
                return False
            
            # 检查元数据
            metadata = data.get('metadata')
            if metadata is None or not isinstance(metadata, dict) or len(metadata) == 0:
                logger.warning("缺少或无效的元数据")
                return False
            
            return True
            
        except Exception as e:
            logger.error(f"输入验证失败: {e}")
            return False
    
    def handle_service_error(self, error: Exception) -> None:
        """
        处理服务错误
        
        Args:
            error: 异常对象
        """
        error_msg = f"因子分析服务错误: {str(error)}"
        logger.error(error_msg)
        st.error(error_msg)
        
        # 更新错误状态
        self._update_analysis_status(f"分析失败: {str(error)}")
    
    def render(self, st_obj, model_data: Any, metadata: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """
        渲染因子分析组件
        
        Args:
            st_obj: Streamlit对象
            model_data: 模型数据
            metadata: 元数据
            
        Returns:
            分析结果字典或None
        """
        try:
            # 验证输入
            input_data = {'model_data': model_data, 'metadata': metadata}
            if not self.validate_input(input_data):
                st_obj.error("[ERROR] 输入数据验证失败，无法进行因子分析。")
                return None
            
            st_obj.markdown("### [DATA] 因子分析结果")
            
            # 更新分析状态
            self._update_analysis_status("分析中...")
            
            # 渲染PCA结果分析
            self._render_pca_results(st_obj, metadata)
            
            # 渲染R²分析
            self._render_r2_analysis(st_obj, metadata)
            
            # 渲染因子载荷热力图
            self._render_factor_loadings_heatmap(st_obj, metadata)
            
            # 渲染因子时间序列
            self._render_factor_time_series(st_obj, metadata)
            
            # 渲染下载按钮
            self._render_download_buttons(st_obj, metadata)
            
            # 更新分析状态
            self._update_analysis_status("分析完成")
            
            # 返回分析结果
            return {
                'pca_results': metadata.get('pca_results_df'),
                'factor_loadings': metadata.get('factor_loadings_df'),
                'factor_series': metadata.get('factor_series'),
                'industry_r2': metadata.get('industry_r2_results'),
                'factor_industry_r2': metadata.get('factor_industry_r2_results'),
                'analysis_status': self._get_analysis_status()
            }
                
        except Exception as e:
            self.handle_service_error(e)
            return None
    
    def _render_pca_results(self, st_obj, metadata: Dict[str, Any]) -> None:
        """
        渲染PCA结果分析
        
        Args:
            st_obj: Streamlit对象
            metadata: 元数据
        """
        st_obj.markdown("**PCA结果分析**")
        
        pca_results = metadata.get('pca_results_df')
        k_factors = self._extract_factor_count(metadata)
        
        if pca_results is not None and isinstance(pca_results, pd.DataFrame) and not pca_results.empty:
            # 格式化PCA结果用于显示
            pca_display = self._format_pca_results_for_display(pca_results, k_factors)
            st_obj.dataframe(pca_display, use_container_width=True)
        else:
            st_obj.write("未找到 PCA 结果。")
    
    def _render_r2_analysis(self, st_obj, metadata: Dict[str, Any]) -> None:
        """
        渲染R²分析
        
        Args:
            st_obj: Streamlit对象
            metadata: 元数据
        """
        st_obj.markdown("---")
        st_obj.markdown("**R² 分析**")
        
        # 创建两列布局
        r2_col1, r2_col2 = st_obj.columns(2)
        
        # 整体R²分析
        with r2_col1:
            st_obj.markdown("**整体 R² (按行业)**")
            industry_r2 = metadata.get('industry_r2_results')
            
            if industry_r2 is not None and isinstance(industry_r2, pd.Series) and not industry_r2.empty:
                st_obj.dataframe(
                    industry_r2.to_frame(name="Industry R2 (All Factors)"), 
                    use_container_width=True
                )
                st_obj.caption(
                    "附注：衡量所有因子共同解释该行业内所有变量整体变动的百分比。"
                    "计算方式为对行业内各变量分别对所有因子进行OLS回归后，"
                    "汇总各变量的总平方和(TSS)与残差平方和(RSS)，"
                    "计算 Pooled R² = 1 - (Sum(RSS) / Sum(TSS))。"
                )
            else:
                st_obj.write("未找到行业整体 R² 数据。")
        
        # 因子对行业R²分析
        with r2_col2:
            st_obj.markdown("**因子对行业 Pooled R²**")
            factor_industry_r2 = metadata.get('factor_industry_r2_results')
            
            if factor_industry_r2 and isinstance(factor_industry_r2, dict):
                try:
                    factor_industry_df = pd.DataFrame(factor_industry_r2)
                    st_obj.dataframe(factor_industry_df, use_container_width=True)
                    st_obj.caption(
                        "附注：衡量单个因子解释该行业内所有变量整体变动的百分比。"
                        "计算方式为对行业内各变量分别对单个因子进行OLS回归后，"
                        "汇总TSS与RSS，计算 Pooled R² = 1 - (Sum(RSS) / Sum(TSS))。"
                    )
                except ValueError as ve:
                    st_obj.warning(f"因子对行业 Pooled R² 数据格式错误，无法转换为DataFrame: {ve}")
                    logger.warning(f"Error converting factor_industry_r2 to DataFrame: {factor_industry_r2}")
            else:
                if factor_industry_r2 is not None and not isinstance(factor_industry_r2, dict):
                    st_obj.write("因子对行业 Pooled R² 数据格式不正确或为空。")
                else:
                    st_obj.write("未找到因子对行业 Pooled R² 数据。")
        
        # 添加数据下载按钮
        st_obj.markdown("---")
        
        # 检查是否有可下载的数据
        has_industry_r2 = industry_r2 is not None and isinstance(industry_r2, pd.Series) and not industry_r2.empty
        has_factor_industry_r2 = (factor_industry_r2 and isinstance(factor_industry_r2, dict) and 
                                 len(factor_industry_r2) > 0)
        
        if has_industry_r2 or has_factor_industry_r2:
            if st_obj.button("数据下载", key="r2_analysis_download_btn", type="secondary"):
                try:
                    excel_data = self._generate_r2_excel(industry_r2, factor_industry_r2)
                    if excel_data:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        filename = f"R2_Analysis_Data_{timestamp}.xlsx"
                        
                        st_obj.download_button(
                            label="下载 R² 分析数据",
                            data=excel_data,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="r2_analysis_download_file"
                        )
                    else:
                        st_obj.error("生成Excel文件失败")
                except Exception as e:
                    st_obj.error(f"生成下载文件时出错: {str(e)}")
                    logger.error(f"Error generating R2 analysis Excel file: {e}")
        else:
            st_obj.info("没有可下载的R²分析数据")
    
    def _render_factor_loadings_heatmap(self, st_obj, metadata: Dict[str, Any]) -> None:
        """
        渲染因子载荷热力图
        
        Args:
            st_obj: Streamlit对象
            metadata: 元数据
        """
        st_obj.markdown("---")
        st_obj.markdown("**因子载荷热力图**")
        
        factor_loadings_df = metadata.get('factor_loadings_df')
        
        if factor_loadings_df is not None and isinstance(factor_loadings_df, pd.DataFrame) and not factor_loadings_df.empty:
            # 执行变量聚类
            clustered_loadings, clustering_success = self._perform_variable_clustering(factor_loadings_df)
            
            # 创建热力图
            fig = self._create_heatmap_figure(
                clustered_loadings, 
                "因子载荷聚类热力图 (Factor Loadings Clustermap)"
            )
            
            if fig is not None:
                st_obj.plotly_chart(fig, use_container_width=True)
                
                if clustering_success:
                    st_obj.info("[SUCCESS] 变量聚类成功，热力图按聚类结果排序显示。")
                else:
                    st_obj.info("[INFO] 变量聚类跳过或失败，热力图按原始顺序显示。")
            else:
                st_obj.error("热力图创建失败。")
        
        elif factor_loadings_df is not None and not isinstance(factor_loadings_df, pd.DataFrame):
            st_obj.warning("因子载荷数据 (factor_loadings_df) 存在但不是有效的 DataFrame 格式。")
        else:
            st_obj.write("未在元数据中找到因子载荷数据 (expected key: 'factor_loadings_df')。")

    def _render_factor_time_series(self, st_obj, metadata: Dict[str, Any]) -> None:
        """
        渲染因子时间序列

        Args:
            st_obj: Streamlit对象
            metadata: 元数据
        """
        st_obj.markdown("---")
        st_obj.markdown("**因子时间序列演变图**")

        factor_series_data = metadata.get('factor_series')

        if factor_series_data is not None and isinstance(factor_series_data, pd.DataFrame) and not factor_series_data.empty:
            factor_names = factor_series_data.columns.tolist()
            num_factors = len(factor_names)

            if num_factors > 0:
                # 计算布局参数
                cols_per_row, num_rows = self._calculate_layout_parameters(num_factors)

                # 为每个因子创建时间序列图
                for row in range(num_rows):
                    # 创建列布局
                    cols = st_obj.columns(cols_per_row)

                    for col_idx in range(cols_per_row):
                        factor_idx = row * cols_per_row + col_idx

                        if factor_idx < num_factors:
                            factor_name = factor_names[factor_idx]

                            with cols[col_idx]:
                                # 创建单个因子的时间序列图
                                factor_data = factor_series_data[factor_name].dropna()

                                if not factor_data.empty:
                                    fig = self._create_factor_time_series_figure(factor_data, factor_name)
                                    if fig is not None:
                                        st_obj.plotly_chart(fig, use_container_width=True)
                                    else:
                                        st_obj.error(f"{factor_name}图表创建失败。")
                                else:
                                    st_obj.warning(f"{factor_name}数据为空，无法绘制图表。")
            else:
                st_obj.write("未找到有效的因子数据。")
        else:
            st_obj.write("未在元数据中找到因子时间序列数据。预期的键名: 'factor_series'。")

    def _render_download_buttons(self, st_obj, metadata: Dict[str, Any]) -> None:
        """
        渲染下载按钮

        Args:
            st_obj: Streamlit对象
            metadata: 元数据
        """
        # 因子载荷数据下载
        factor_loadings_df = metadata.get('factor_loadings_df')
        if factor_loadings_df is not None and isinstance(factor_loadings_df, pd.DataFrame) and not factor_loadings_df.empty:
            try:
                csv_loadings = factor_loadings_df.to_csv(index=True).encode('utf-8-sig')
                st_obj.download_button(
                    label="下载因子载荷数据",
                    data=csv_loadings,
                    file_name="factor_loadings.csv",
                    mime="text/csv",
                    key="download_factor_loadings"
                )
            except Exception as e:
                st_obj.error(f"生成因子载荷下载文件时出错: {e}")

        # 因子时间序列数据下载
        factor_series_data = metadata.get('factor_series')
        if factor_series_data is not None and isinstance(factor_series_data, pd.DataFrame) and not factor_series_data.empty:
            try:
                all_factors_csv = factor_series_data.to_csv(index=True).encode('utf-8-sig')
                st_obj.download_button(
                    label="下载因子时间序列数据",
                    data=all_factors_csv,
                    file_name="所有因子时间序列.csv",
                    mime="text/csv",
                    key="download_all_factors_timeseries"
                )
            except Exception as e:
                st_obj.error(f"生成所有因子下载文件时出错: {e}")

    def _extract_factor_count(self, metadata: Dict[str, Any]) -> int:
        """
        提取因子数量

        Args:
            metadata: 元数据

        Returns:
            因子数量
        """
        # 尝试从best_params获取
        k = metadata.get('best_params', {}).get('k_factors', metadata.get('n_factors', 0))

        if not isinstance(k, int) or k <= 0:
            # 从PCA数据推断
            pca_results = metadata.get('pca_results_df')
            if pca_results is not None and isinstance(pca_results, pd.DataFrame):
                k = len(pca_results.index)
            else:
                k = 0

        return k

    def _format_pca_results_for_display(self, pca_results: pd.DataFrame, k_factors: int) -> pd.DataFrame:
        """
        格式化PCA结果用于显示

        Args:
            pca_results: PCA结果DataFrame
            k_factors: 因子数量

        Returns:
            格式化后的DataFrame
        """
        # 选择前k个因子
        pca_display = pca_results.head(k_factors if k_factors > 0 else len(pca_results.index)).copy()

        # 移除旧的主成分列（如果存在）
        if '主成分 (PC)' in pca_display.columns:
            pca_display = pca_display.drop(columns=['主成分 (PC)'])

        # 添加新的主成分列
        pca_display.insert(0, '主成分 (PC)', [f"PC{i+1}" for i in range(len(pca_display.index))])

        # 重置索引
        if not isinstance(pca_display.index, pd.RangeIndex):
            pca_display = pca_display.reset_index()
            if 'index' in pca_display.columns:
                pca_display = pca_display.rename(columns={'index': 'Original Index'})

        # 重命名列
        pca_display = pca_display.rename(columns={
            '解释方差 (%)': '解释方差(%)',
            '累计解释方差 (%)': '累计解释方差(%)',
            '特征值 (Eigenvalue)': '特征值(Eigenvalue)'
        })

        return pca_display

    def _perform_variable_clustering(self, loadings_df: pd.DataFrame) -> Tuple[pd.DataFrame, bool]:
        """
        执行变量聚类

        Args:
            loadings_df: 因子载荷DataFrame

        Returns:
            (聚类后的DataFrame, 聚类是否成功)
        """
        if loadings_df.shape[0] <= 1:
            logger.info("因子载荷热力图：只有一个变量，跳过聚类。")
            return loadings_df, False

        try:
            # 执行层次聚类
            linked = sch.linkage(loadings_df.values, method='ward', metric='euclidean')
            dendro = sch.dendrogram(linked, no_plot=True)
            clustered_indices = dendro['leaves']

            # 重新排序
            clustered_loadings = loadings_df.iloc[clustered_indices, :]

            logger.info("因子载荷热力图：变量聚类成功。")
            return clustered_loadings, True

        except Exception as e:
            logger.warning(f"因子载荷热力图的变量聚类失败: {e}")
            return loadings_df, False

    def _create_heatmap_figure(self, loadings_df: pd.DataFrame, title: str) -> Optional[go.Figure]:
        """
        创建热力图

        Args:
            loadings_df: 因子载荷DataFrame
            title: 图表标题

        Returns:
            Plotly图表对象或None
        """
        try:
            # 准备数据
            z_values = loadings_df.values
            x_labels = loadings_df.columns.tolist()  # 因子名
            y_labels = loadings_df.index.tolist()    # 变量名

            # 创建热力图
            fig = go.Figure(data=go.Heatmap(
                z=z_values,
                x=x_labels,
                y=y_labels,
                colorscale=self._default_chart_settings['heatmap_colorscale'],
                zmid=0,
                colorbar=dict(title='载荷值'),
                xgap=1,
                ygap=1,
                hovertemplate=(
                    "变量 (Variable): %{y}<br>" +
                    "因子 (Factor): %{x}<br>" +
                    "载荷值 (Loading): %{z:.4f}<extra></extra>"
                )
            ))

            # 添加文本注释
            annotations = []
            for i, var_name in enumerate(y_labels):
                for j, factor_name in enumerate(x_labels):
                    val = z_values[i][j]
                    annotations.append(
                        go.layout.Annotation(
                            text=f"{val:.2f}",
                            x=factor_name,
                            y=var_name,
                            xref='x1',
                            yref='y1',
                            showarrow=False,
                            font=dict(color='white' if abs(val) > 0.5 else 'black')
                        )
                    )

            # 更新布局
            fig.update_layout(
                title=title,
                xaxis_title="因子 (Factors)",
                yaxis_title="变量 (Predictors)",
                yaxis=dict(
                    type='category',
                    categoryorder='array',
                    categoryarray=y_labels
                ),
                height=max(self._default_chart_settings['heatmap_height'], len(y_labels) * 20),
                margin=dict(l=150, r=30, t=80, b=100),
                annotations=annotations
            )

            fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='LightGrey')
            fig.update_xaxes(showgrid=False)

            return fig

        except Exception as e:
            logger.error(f"创建热力图失败: {e}")
            return None

    def _create_factor_time_series_figure(self, factor_data: pd.Series, factor_name: str) -> Optional[go.Figure]:
        """
        创建因子时间序列图

        Args:
            factor_data: 因子时间序列数据
            factor_name: 因子名称

        Returns:
            Plotly图表对象或None
        """
        try:
            fig = go.Figure()

            fig.add_trace(go.Scatter(
                x=factor_data.index,
                y=factor_data.values,
                mode='lines+markers',
                name=factor_name,
                line=dict(width=2),
                marker=dict(size=4),
                hovertemplate=(
                    f"日期: %{{x|%Y/%m/%d}}<br>" +
                    f"{factor_name}: %{{y:.4f}}<extra></extra>"
                )
            ))

            fig.update_layout(
                title=f"{factor_name}",
                xaxis_title="日期",
                yaxis_title="因子值",
                height=self._default_chart_settings['time_series_height'],
                margin=dict(t=60, b=80, l=60, r=30),
                showlegend=False,
                hovermode='x unified'
            )

            # 添加零轴线
            fig.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)

            return fig

        except Exception as e:
            logger.error(f"创建因子时间序列图失败: {e}")
            return None

    def _calculate_layout_parameters(self, num_factors: int) -> Tuple[int, int]:
        """
        计算布局参数

        Args:
            num_factors: 因子数量

        Returns:
            (每行列数, 行数)
        """
        try:
            # 尝试从配置获取
            cols_per_row = (VisualizationDefaults.FACTOR_PLOT_COLS_EVEN
                          if num_factors % 2 == 0
                          else VisualizationDefaults.FACTOR_PLOT_COLS_ODD)
        except Exception:
            # 使用默认值
            cols_per_row = 2 if num_factors % 2 == 0 else 3

        num_rows = (num_factors + cols_per_row - 1) // cols_per_row

        return cols_per_row, num_rows

    def _get_analysis_status(self) -> str:
        """
        获取当前分析状态

        Returns:
            分析状态字符串
        """
        return self._get_state('dfm_analysis_status', '等待分析')

    def _update_analysis_status(self, status: str) -> None:
        """
        更新分析状态

        Args:
            status: 新的分析状态
        """
        self._set_state('dfm_analysis_status', status)

    def _get_state(self, key: str, default: Any = None) -> Any:
        """获取状态值"""
        try:
            dfm_manager = get_global_dfm_manager()
            if dfm_manager:
                value = dfm_manager.get_dfm_state('model_analysis', key, None)
                if value is not None:
                    return value

            # 备用方案：使用统一状态管理器
            from dashboard.core import get_unified_manager
            state_manager = get_unified_manager()
            if state_manager:
                backup_key = f'dfm_model_analysis_{key}'
                return state_manager.get_state(backup_key, default)
            return default

        except Exception as e:
            logger.warning(f"获取状态失败: {e}")
            return default

    def _set_state(self, key: str, value: Any) -> None:
        """设置状态值"""
        try:
            dfm_manager = get_global_dfm_manager()
            if dfm_manager:
                success = dfm_manager.set_dfm_state('model_analysis', key, value)
                if not success:
                    raise RuntimeError(f"Failed to set state via DFM manager: {key}")
            else:
                raise RuntimeError("DFM manager not available")

        except Exception as e:
            logger.error(f"设置状态失败: {e}")
            raise
    
    def _generate_r2_excel(self, industry_r2: pd.Series, factor_industry_r2: dict) -> bytes:
        """
        生成R²分析数据的Excel文件
        
        Args:
            industry_r2: 整体R²数据(按行业)
            factor_industry_r2: 因子对行业Pooled R²数据
            
        Returns:
            bytes: Excel文件的字节数据
        """
        try:
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "R2_Analysis"
            
            current_row = 1
            
            # 添加标题
            ws.cell(row=current_row, column=1, value="R² 分析数据报告")
            ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=14)
            current_row += 2
            
            ws.cell(row=current_row, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            current_row += 3
            
            # 第一个表：整体 R² (按行业)
            if industry_r2 is not None and isinstance(industry_r2, pd.Series) and not industry_r2.empty:
                ws.cell(row=current_row, column=1, value="整体 R² (按行业)")
                ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=12)
                current_row += 2
                
                # 添加表头
                ws.cell(row=current_row, column=1, value="行业")
                ws.cell(row=current_row, column=2, value="Industry R2 (All Factors)")
                ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True)
                ws.cell(row=current_row, column=2).font = ws.cell(row=current_row, column=2).font.copy(bold=True)
                current_row += 1
                
                # 添加数据
                for index, value in industry_r2.items():
                    ws.cell(row=current_row, column=1, value=str(index))
                    ws.cell(row=current_row, column=2, value=float(value) if pd.notna(value) else None)
                    current_row += 1
                
                # 添加附注
                current_row += 1
                ws.cell(row=current_row, column=1, value="附注：")
                ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True)
                current_row += 1
                ws.cell(row=current_row, column=1, value="衡量所有因子共同解释该行业内所有变量整体变动的百分比。")
                current_row += 1
                ws.cell(row=current_row, column=1, value="计算方式为对行业内各变量分别对所有因子进行OLS回归后，")
                current_row += 1
                ws.cell(row=current_row, column=1, value="汇总各变量的总平方和(TSS)与残差平方和(RSS)，")
                current_row += 1
                ws.cell(row=current_row, column=1, value="计算 Pooled R² = 1 - (Sum(RSS) / Sum(TSS))。")
                current_row += 3
            
            # 第二个表：因子对行业 Pooled R²
            if factor_industry_r2 and isinstance(factor_industry_r2, dict) and len(factor_industry_r2) > 0:
                try:
                    factor_industry_df = pd.DataFrame(factor_industry_r2)
                    
                    ws.cell(row=current_row, column=1, value="因子对行业 Pooled R²")
                    ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=12)
                    current_row += 2
                    
                    # 添加表头
                    col_offset = 0
                    for col_idx, column in enumerate(['行业/因子'] + list(factor_industry_df.columns)):
                        ws.cell(row=current_row, column=col_idx + 1, value=column)
                        ws.cell(row=current_row, column=col_idx + 1).font = ws.cell(row=current_row, column=col_idx + 1).font.copy(bold=True)
                    current_row += 1
                    
                    # 添加数据
                    for row_idx, (index, row_data) in enumerate(factor_industry_df.iterrows()):
                        ws.cell(row=current_row, column=1, value=str(index))
                        for col_idx, value in enumerate(row_data):
                            ws.cell(row=current_row, column=col_idx + 2, value=float(value) if pd.notna(value) else None)
                        current_row += 1
                    
                    # 添加附注
                    current_row += 1
                    ws.cell(row=current_row, column=1, value="附注：")
                    ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True)
                    current_row += 1
                    ws.cell(row=current_row, column=1, value="衡量单个因子解释该行业内所有变量整体变动的百分比。")
                    current_row += 1
                    ws.cell(row=current_row, column=1, value="计算方式为对行业内各变量分别对单个因子进行OLS回归后，")
                    current_row += 1
                    ws.cell(row=current_row, column=1, value="汇总TSS与RSS，计算 Pooled R² = 1 - (Sum(RSS) / Sum(TSS))。")
                    
                except Exception as e:
                    logger.error(f"Error processing factor_industry_r2 data for Excel: {e}")
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存到字节流
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating R2 Excel file: {e}")
            return None
