# -*- coding: utf-8 -*-
"""
DFM模型分析页面组件

完全重构版本，与old_ui/dfm_ui.py保持完全一致
"""

import pandas as pd
import logging
import plotly.graph_objects as go
import numpy as np
import joblib
import pickle
import sys
import os
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 统一状态管理器导入
current_dir = os.path.dirname(os.path.abspath(__file__))
dashboard_root = os.path.abspath(os.path.join(current_dir, '..', '..', '..', '..'))
if dashboard_root not in sys.path:
    sys.path.insert(0, dashboard_root)

from dashboard.core import get_unified_manager
print("[DFM Model Analysis] [SUCCESS] 统一状态管理器导入成功")


def get_dfm_state(key, default=None):
    """获取DFM状态值"""
    unified_manager = get_unified_manager()
    if not unified_manager:
        return default

    # 简化：数据相关的键从data_prep命名空间获取，其他从model_analysis获取
    data_keys = [
        'dfm_prepared_data_df',
        'dfm_transform_log_obj',
        'dfm_industry_map_obj',
        'dfm_removed_vars_log_obj',
        'dfm_var_type_map_obj'
    ]

    if key in data_keys:
        return unified_manager.get_dfm_state('data_prep', key, default)
    else:
        return unified_manager.get_dfm_state('model_analysis', key, default)


def set_dfm_state(key, value):
    """设置DFM状态值"""
    unified_manager = get_unified_manager()
    if not unified_manager:
        print(f"[DFM Model Analysis] [WARNING] 统一状态管理器不可用: {key}")
        return False

    success = unified_manager.set_dfm_state('model_analysis', key, value)
    if not success:
        print(f"[DFM Model Analysis] [WARNING] 统一状态管理器设置失败: {key}")
    return success

from typing import Optional, Dict, Any
from datetime import datetime

# 配置已移除，所有参数通过UI设置
CONFIG_AVAILABLE = False

# Import backend functions
from dashboard.models.DFM.results.dfm_backend import (
    load_dfm_results_from_uploads,
    perform_loadings_clustering
)

logger = logging.getLogger(__name__)


def generate_r2_excel(industry_r2: pd.Series, factor_industry_r2: dict) -> bytes:
    """
    生成R²分析数据的Excel文件
    
    Args:
        industry_r2: 整体R²数据(按行业)
        factor_industry_r2: 因子对行业Pooled R²数据
        
    Returns:
        bytes: Excel文件的字节数据
    """
    try:
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "R2_Analysis"
        
        current_row = 1
        
        # 添加标题
        ws.cell(row=current_row, column=1, value="R² 分析数据报告")
        ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=14)
        current_row += 2
        
        ws.cell(row=current_row, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        current_row += 3
        
        # 第一个表：整体 R² (按行业)
        if industry_r2 is not None and isinstance(industry_r2, pd.Series) and not industry_r2.empty:
            ws.cell(row=current_row, column=1, value="整体 R² (按行业)")
            ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=12)
            current_row += 2
            
            # 添加表头
            ws.cell(row=current_row, column=1, value="行业")
            ws.cell(row=current_row, column=2, value="Industry R2 (All Factors)")
            ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True)
            ws.cell(row=current_row, column=2).font = ws.cell(row=current_row, column=2).font.copy(bold=True)
            current_row += 1
            
            # 添加数据
            for index, value in industry_r2.items():
                ws.cell(row=current_row, column=1, value=str(index))
                ws.cell(row=current_row, column=2, value=float(value) if pd.notna(value) else None)
                current_row += 1
            
            # 添加附注
            current_row += 1
            ws.cell(row=current_row, column=1, value="附注：")
            ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True)
            current_row += 1
            ws.cell(row=current_row, column=1, value="衡量所有因子共同解释该行业内所有变量整体变动的百分比。")
            current_row += 1
            ws.cell(row=current_row, column=1, value="计算方式为对行业内各变量分别对所有因子进行OLS回归后，")
            current_row += 1
            ws.cell(row=current_row, column=1, value="汇总各变量的总平方和(TSS)与残差平方和(RSS)，")
            current_row += 1
            ws.cell(row=current_row, column=1, value="计算 Pooled R² = 1 - (Sum(RSS) / Sum(TSS))。")
            current_row += 3
        
        # 第二个表：因子对行业 Pooled R²
        if factor_industry_r2 and isinstance(factor_industry_r2, dict) and len(factor_industry_r2) > 0:
            try:
                factor_industry_df = pd.DataFrame(factor_industry_r2)
                
                ws.cell(row=current_row, column=1, value="因子对行业 Pooled R²")
                ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=12)
                current_row += 2
                
                # 添加表头
                col_offset = 0
                for col_idx, column in enumerate(['行业/因子'] + list(factor_industry_df.columns)):
                    ws.cell(row=current_row, column=col_idx + 1, value=column)
                    ws.cell(row=current_row, column=col_idx + 1).font = ws.cell(row=current_row, column=col_idx + 1).font.copy(bold=True)
                current_row += 1
                
                # 添加数据
                for row_idx, (index, row_data) in enumerate(factor_industry_df.iterrows()):
                    ws.cell(row=current_row, column=1, value=str(index))
                    for col_idx, value in enumerate(row_data):
                        ws.cell(row=current_row, column=col_idx + 2, value=float(value) if pd.notna(value) else None)
                    current_row += 1
                
                # 添加附注
                current_row += 1
                ws.cell(row=current_row, column=1, value="附注：")
                ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True)
                current_row += 1
                ws.cell(row=current_row, column=1, value="衡量单个因子解释该行业内所有变量整体变动的百分比。")
                current_row += 1
                ws.cell(row=current_row, column=1, value="计算方式为对行业内各变量分别对单个因子进行OLS回归后，")
                current_row += 1
                ws.cell(row=current_row, column=1, value="汇总TSS与RSS，计算 Pooled R² = 1 - (Sum(RSS) / Sum(TSS))。")
                
            except Exception as e:
                logger.error(f"Error processing factor_industry_r2 data for Excel: {e}")
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating R2 Excel file: {e}")
        return None


def load_dfm_data() -> tuple[Optional[Any], Optional[Dict]]:
    """从 session_state 加载模型结果和元数据。"""
    model_file = get_dfm_state('dfm_model_file_indep', None)
    metadata_file = get_dfm_state('dfm_metadata_file_indep', None)

    model_results = None
    metadata = None

    # 安全地检查模型文件是否存在且为有效的文件对象
    if model_file is not None:
        # 增加文件对象类型检查，避免对非文件对象调用seek方法
        if hasattr(model_file, 'seek') and hasattr(model_file, 'read'):
            try:
                model_file.seek(0) # 重置文件指针
                model_results = joblib.load(model_file)
                print("[DFM UI] Model loaded successfully from session state.")
            except Exception as e:
                import streamlit as st
                file_name = getattr(model_file, 'name', '未知文件')
                st.error(f"加载模型文件 ('{file_name}') 时出错: {e}")
        else:
            print(f"[DFM UI] 检测到无效的模型文件对象类型: {type(model_file)}")

    # 安全地检查元数据文件是否存在且为有效的文件对象
    if metadata_file is not None:
        # 增加文件对象类型检查，避免对非文件对象调用seek方法
        if hasattr(metadata_file, 'seek') and hasattr(metadata_file, 'read'):
            try:
                metadata_file.seek(0) # 重置文件指针
                metadata = pickle.load(metadata_file)
                print("[DFM UI] Metadata loaded successfully from session state.")
            except Exception as e:
                import streamlit as st
                file_name = getattr(metadata_file, 'name', '未知文件')
                st.error(f"加载元数据文件 ('{file_name}') 时出错: {e}")
        else:
            print(f"[DFM UI] 检测到无效的元数据文件对象类型: {type(metadata_file)}")

    return model_results, metadata

def plot_factor_evolution(factor_df: pd.DataFrame, title: str = "因子时间序列演变图"):
    """绘制因子随时间变化的曲线图。"""
    import streamlit as st  # 懒加载导入

    if not isinstance(factor_df, pd.DataFrame) or factor_df.empty:
        st.warning("因子数据无效，无法绘制演变图。")
        return
    
    fig = go.Figure()
    
    for col in factor_df.columns:
        fig.add_trace(go.Scatter(
            x=factor_df.index,
            y=factor_df[col],
            mode='lines',
            name=col,
            hovertemplate=(
                f"日期: %{{x}}<br>" +
                f"{col}: %{{y:.4f}}<extra></extra>"
            )
        ))

    fig.update_layout(
        title=title,
        xaxis_title="日期",
        yaxis_title="因子值",
        legend_title_text='因子',
        hovermode='x unified',
        height=400,
        margin=dict(t=50, b=50, l=50, r=50)
    )

    st.plotly_chart(fig, use_container_width=True)

def plot_loadings_heatmap(loadings_df: pd.DataFrame, title: str = "因子载荷矩阵 (Lambda)", cluster_vars: bool = True):
    """
    绘制因子载荷矩阵的热力图 (因子在X轴, 变量在Y轴, 可选聚类)。

    Args:
        loadings_df: 包含因子载荷的 DataFrame (原始形式：变量为行，因子为列)。
        title: 图表标题。
        cluster_vars: 是否对变量进行聚类排序。
    """
    import streamlit as st  # 懒加载导入

    if not isinstance(loadings_df, pd.DataFrame) or loadings_df.empty:
        st.warning(f"无法绘制热力图：提供的载荷数据无效 ({title})。")
        return

    # 使用Backend的聚类函数
    data_for_clustering, variable_names, clustering_success = perform_loadings_clustering(
        loadings_df,
        cluster_vars=cluster_vars
    )

    if clustering_success:
        title += " (变量聚类排序)"

    factor_names = data_for_clustering.columns.tolist()

    # 2. 转置数据以便绘图 (因子在 X 轴, 变量在 Y 轴)
    plot_data_transposed = data_for_clustering.T # 转置后：因子是行，（聚类后）变量是列
    
    # 确保轴标签列表是最新的
    y_axis_labels = variable_names # 聚类后的变量名
    x_axis_labels = factor_names   # 原始因子名

    # 3. 创建热力图
    fig = go.Figure(data=go.Heatmap(
        z=plot_data_transposed.values, # 使用转置后的数据
        x=x_axis_labels,          # X 轴是因子
        y=y_axis_labels,          # Y 轴是变量 (按聚类顺序)
        colorscale='RdBu',
        zmid=0,
        hovertemplate=(
            "变量 (Variable): %{y}<br>" +
            "因子 (Factor): %{x}<br>" +
            "载荷值 (Loading): %{z:.4f}<extra></extra>"
        )
    ))

    # 4. 更新布局
    fig.update_layout(
        title=title,
        xaxis_title="因子 (Factors)",
        yaxis_title="变量 (Predictors)",
        xaxis_tickangle=-45, 
        # Y轴使用类别类型，并直接指定顺序
        yaxis=dict(
            type='category', 
            categoryorder='array', # 明确指定使用下面提供的数组顺序
            categoryarray=y_axis_labels # 确保Y轴按聚类顺序显示
        ), 
        height=max(600, len(y_axis_labels) * 20), # 调整高度计算
        margin=dict(l=150, r=30, t=80, b=100) # 减小左右边距
    )
    
    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='LightGrey')
    fig.update_xaxes(showgrid=False) 

    st.plotly_chart(fig, use_container_width=True)

def render_file_upload_section(st_instance):
    """
    渲染文件上传区域
    """
    def cleanup_invalid_file_states():
        """清理可能存在的无效文件状态"""
        model_file = get_dfm_state('dfm_model_file_indep', None)
        metadata_file = get_dfm_state('dfm_metadata_file_indep', None)
        
        # 检查模型文件状态
        if model_file is not None:
            if not (hasattr(model_file, 'seek') and 
                   hasattr(model_file, 'read') and 
                   hasattr(model_file, 'name') and
                   getattr(model_file, 'name', '未知文件') != '未知文件'):
                set_dfm_state('dfm_model_file_indep', None)
        
        # 检查元数据文件状态
        if metadata_file is not None:
            if not (hasattr(metadata_file, 'seek') and 
                   hasattr(metadata_file, 'read') and 
                   hasattr(metadata_file, 'name') and
                   getattr(metadata_file, 'name', '未知文件') != '未知文件'):
                set_dfm_state('dfm_metadata_file_indep', None)
    
    # 执行状态清理
    cleanup_invalid_file_states()

    st_instance.markdown("### 模型文件上传")

    # 创建两列布局
    col_model, col_metadata = st_instance.columns(2)  
    
    with col_model:
        st_instance.markdown("**DFM 模型文件 (.joblib)**")
        uploaded_model_file = st_instance.file_uploader(
            "选择模型文件",
            type=['joblib'],
            key="dfm_model_upload_independent",
            help="上传训练好的DFM模型文件，通常为.joblib格式"
        )
        
        if uploaded_model_file:
            set_dfm_state("dfm_model_file_indep", uploaded_model_file)
        else:
            existing_model_file = get_dfm_state('dfm_model_file_indep', None)
            if existing_model_file is not None:
                # 修复：只在有效文件时显示文件名
                file_name = getattr(existing_model_file, 'name', '未知文件')
                if (hasattr(existing_model_file, 'seek') and 
                    hasattr(existing_model_file, 'read') and 
                    file_name != '未知文件'):
                    st_instance.info(f"当前文件: {file_name}")
    
    with col_metadata:
        st_instance.markdown("**元数据文件 (.pkl)**")
        uploaded_metadata_file = st_instance.file_uploader(
            "选择元数据文件", 
            type=['pkl'],
            key="dfm_metadata_upload_independent",
            help="上传包含训练元数据的.pkl文件"
        )
        
        if uploaded_metadata_file:
            set_dfm_state("dfm_metadata_file_indep", uploaded_metadata_file)
        else:
            existing_metadata_file = get_dfm_state('dfm_metadata_file_indep', None)
            if existing_metadata_file is not None:
                # 修复：只在有效文件时显示文件名
                file_name = getattr(existing_metadata_file, 'name', '未知文件')
                if (hasattr(existing_metadata_file, 'seek') and 
                    hasattr(existing_metadata_file, 'read') and 
                    file_name != '未知文件'):
                    st_instance.info(f"当前文件: {file_name}")
    
    # 文件状态总结
    model_file = get_dfm_state('dfm_model_file_indep', None)
    metadata_file = get_dfm_state('dfm_metadata_file_indep', None)

    # 修复：严格检查文件是否真正有效
    def is_valid_file_object(file_obj):
        """检查是否为有效的文件对象"""
        if file_obj is None:
            return False
        # 检查是否具有文件对象的必要方法
        return (hasattr(file_obj, 'seek') and 
                hasattr(file_obj, 'read') and 
                hasattr(file_obj, 'name') and
                file_obj.name != '未知文件')

    model_file_exists = is_valid_file_object(model_file)
    metadata_file_exists = is_valid_file_object(metadata_file)

    if model_file_exists and metadata_file_exists:
        return True
    else:
        missing_files = []
        if not model_file_exists:
            missing_files.append("模型文件")
        if not metadata_file_exists:
            missing_files.append("元数据文件")

        st_instance.warning(f"[WARNING] 缺少文件: {', '.join(missing_files)}。请上传所有文件后再进行分析。")
        return False

def render_dfm_tab(st):
    """Renders the DFM Model Results tab using independently uploaded files."""
        
    # 添加文件上传区域
    files_ready = render_file_upload_section(st)
    
    if not files_ready:
        st.info("[INFO] 请先上传模型文件和元数据文件以继续分析。")
        return

    # 修复：只有在文件准备好后才尝试加载数据
    try:
        model_results, metadata = load_dfm_data()
    except Exception as e:
        st.error(f"[ERROR] 加载文件时出现意外错误: {e}")
        return

    if model_results is None or metadata is None:
        st.error("[ERROR] 无法加载模型数据，请检查文件格式和内容。")
        return

    # 调用后端处理函数
    model, metadata, load_errors = load_dfm_results_from_uploads(model_results, metadata)
    
    all_errors = load_errors

    if all_errors:
        st.error("加载 DFM 相关文件时遇到错误:")
        for error in all_errors:
            st.error(f"- {error}")
        if model is None or metadata is None: 
            return
            
    if model is None or metadata is None:
        st.warning("未能成功加载 DFM 模型或元数据。请检查文件内容或格式。")
        return

    # 获取验证期信息（用于后续绘图标记，不显示）
    val_start = metadata.get('validation_start_date', 'N/A')
    val_end = metadata.get('validation_end_date', 'N/A')

    # 也尝试从train_end_date获取训练期结束日期
    train_end = metadata.get('train_end_date', metadata.get('train_end', 'N/A'))

    logger.info(f"[CHART DEBUG] 从metadata获取的日期配置:")
    logger.info(f"[CHART DEBUG]   train_end={train_end}")
    logger.info(f"[CHART DEBUG]   val_start={val_start}")
    logger.info(f"[CHART DEBUG]   val_end={val_end}")

    # 从元数据获取指标
    # 修复因子数量获取逻辑 - 多重回退策略
    k_factors = 'N/A'

    # 策略1: 从best_params获取 (支持多种键名)
    best_params = metadata.get('best_params', {})
    if isinstance(best_params, dict):
        # 尝试多种可能的键名
        possible_keys = ['k_factors', 'k_factors_final', 'best_k_factors']
        for key in possible_keys:
            if key in best_params:
                k_factors = best_params[key]
                logger.info(f"从best_params['{key}']获取k_factors: {k_factors}")
                break

    # 策略2: 从多种可能的因子数键获取
    if k_factors == 'N/A':
        factor_keys = ['n_factors', 'k_factors_final', 'best_k_factors']
        for key in factor_keys:
            value = metadata.get(key)
            if value is not None and value != 'N/A':
                k_factors = value
                logger.info(f"从{key}获取k_factors: {k_factors}")
                break

    # 策略3: 从optimal_k_factors获取
    if k_factors == 'N/A':
        optimal_k = metadata.get('optimal_k_factors')
        if optimal_k is not None and optimal_k != 'N/A':
            k_factors = optimal_k
            logger.info(f"从optimal_k_factors获取k_factors: {k_factors}")

    # 策略4: 从factor_loadings推断
    if k_factors == 'N/A':
        factor_loadings = metadata.get('factor_loadings_df')
        if factor_loadings is not None and hasattr(factor_loadings, 'columns'):
            k_factors = len(factor_loadings.columns)
            logger.info(f"从factor_loadings推断k_factors: {k_factors}")

    # 策略5: 从factor_series推断
    if k_factors == 'N/A':
        factor_series = metadata.get('factor_series')
        if factor_series is not None and hasattr(factor_series, 'columns'):
            k_factors = len(factor_series.columns)
            logger.info(f"从factor_series推断k_factors: {k_factors}")

    # 最终检查
    if k_factors == 'N/A':
        logger.warning("无法获取因子数量，将显示N/A")
    else:
        logger.info(f"最终确定的k_factors: {k_factors}")

    # 修复变量数量获取逻辑
    best_variables = metadata.get('best_variables', [])
    if isinstance(best_variables, list) and len(best_variables) > 0:
        n_vars = len(best_variables)
        logger.info(f"从best_variables获取变量数量: {n_vars}")
    else:
        # 从factor_loadings推断变量数量
        factor_loadings = metadata.get('factor_loadings_df')
        if factor_loadings is not None and hasattr(factor_loadings, 'index'):
            n_vars = len(factor_loadings.index)
            logger.info(f"从factor_loadings推断变量数量: {n_vars}")
        else:
            n_vars = 'N/A'
            logger.warning("无法获取变量数量，将显示N/A")

    revised_is_hr = metadata.get('revised_is_hr')
    revised_oos_hr = metadata.get('revised_oos_hr')
    revised_is_rmse = metadata.get('revised_is_rmse')
    revised_oos_rmse = metadata.get('revised_oos_rmse')
    revised_is_mae = metadata.get('revised_is_mae')
    revised_oos_mae = metadata.get('revised_oos_mae')

    def format_value(val, is_percent=False, precision=2):
        if isinstance(val, (int, float)) and pd.notna(val):
            if is_percent:
                # MODIFIED: Assume val is already the percentage value if is_percent is True
                # e.g., if val is 72.3, it represents 72.3%
                return f"{val:.{precision}f}%" 
            return f"{val:.{precision}f}"
        return 'N/A' if val == 'N/A' or pd.isna(val) else str(val)

    row1_col1, row1_col2 = st.columns(2)
    with row1_col1:
        # 兼容numpy整数类型和Python整数类型
        display_k = int(k_factors) if isinstance(k_factors, (int, np.integer)) else 'N/A'
        st.metric("最终因子数 (k)", display_k)
    with row1_col2:
        # 兼容numpy整数类型和Python整数类型
        display_n = int(n_vars) if isinstance(n_vars, (int, np.integer)) else 'N/A'
        st.metric("最终变量数 (N)", display_n)

    row2_col1, row2_col2 = st.columns(2)
    with row2_col1:
        st.metric("训练期胜率", format_value(revised_is_hr, is_percent=True))
    with row2_col2:
        st.metric("验证期胜率", format_value(revised_oos_hr, is_percent=True))

    row3_col1, row3_col2, row3_col3, row3_col4 = st.columns(4)
    with row3_col1:
        st.metric("样本内 RMSE", format_value(revised_is_rmse))
    with row3_col2:
        st.metric("样本外 RMSE", format_value(revised_oos_rmse))
    with row3_col3:
        st.metric("样本内 MAE", format_value(revised_is_mae))
    with row3_col4:
        st.metric("样本外 MAE", format_value(revised_oos_mae))

    # 修复：直接使用pickle文件中的complete_aligned_table数据
    complete_aligned_table = metadata.get('complete_aligned_table')

    # 获取目标变量名
    target_variable_name_for_plot = metadata.get('target_variable', '规模以上工业增加值:当月同比')

    logger.info("[HOT] 开始验证complete_aligned_table数据...")
    logger.info(f"[HOT] metadata中的键: {list(metadata.keys())}")
    logger.info(f"[HOT] complete_aligned_table类型: {type(complete_aligned_table)}")

    if complete_aligned_table is not None:
        if isinstance(complete_aligned_table, pd.DataFrame):
            logger.info(f"[HOT] complete_aligned_table形状: {complete_aligned_table.shape}")
            logger.info(f"[HOT] complete_aligned_table列名: {list(complete_aligned_table.columns)}")
            logger.info(f"[HOT] complete_aligned_table是否为空: {complete_aligned_table.empty}")
            if not complete_aligned_table.empty:
                logger.info(f"[HOT] complete_aligned_table时间范围: {complete_aligned_table.index.min()} 到 {complete_aligned_table.index.max()}")
                logger.info(f"[HOT] complete_aligned_table非空值统计: {complete_aligned_table.notna().sum().to_dict()}")
        else:
            logger.warning(f"[HOT] complete_aligned_table不是DataFrame，而是: {type(complete_aligned_table)}")
    else:
        logger.error("[HOT] complete_aligned_table为None")

    if complete_aligned_table is not None and isinstance(complete_aligned_table, pd.DataFrame) and not complete_aligned_table.empty:
        # 直接使用pickle文件中的数据
        logger.info("[SUCCESS] 使用pickle文件中的complete_aligned_table数据")
        comparison_df = complete_aligned_table.copy()

        # 确保列名正确
        # 列名：'Nowcast (Original Scale)' 和目标变量名
        nowcast_display_name = "Nowcast值"
        target_display_name = target_variable_name_for_plot

        # 检查并重命名列
        if len(comparison_df.columns) >= 2:
            # 第一列是Nowcast，第二列是Target
            comparison_df.columns = [nowcast_display_name, target_display_name]

        logger.info(f"数据包含 {len(comparison_df)} 行数据")
        logger.info(f"时间范围: {comparison_df.index.min()} 到 {comparison_df.index.max()}")

    else:
        # 如果没有数据，显示错误信息
        logger.error("[ERROR] 未找到complete_aligned_table数据")
        comparison_df = None

    # 检查是否成功获取了对比数据
    if comparison_df is not None and not comparison_df.empty:
        # 确保索引是DatetimeIndex
        if not isinstance(comparison_df.index, pd.DatetimeIndex):
            try:
                comparison_df.index = pd.to_datetime(comparison_df.index)
                comparison_df = comparison_df.sort_index()
            except Exception as e:
                st.error(f"无法将对比数据的索引转换为日期时间格式: {e}")
                st.dataframe(comparison_df, use_container_width=True)
                return

        # 绘制Nowcast vs 实际值图表
        logger.info("开始绘制 Nowcast vs 实际值图表...")
        fig = go.Figure()

        # 添加Nowcast数据线
        if nowcast_display_name in comparison_df.columns and comparison_df[nowcast_display_name].notna().any():
            fig.add_trace(go.Scatter(
                x=comparison_df.index,
                y=comparison_df[nowcast_display_name],
                mode='lines+markers',
                name=nowcast_display_name,
                line=dict(color='blue'),
                marker=dict(size=5),
                hovertemplate=
                f'<b>日期</b>: %{{x|%Y/%m/%d}}<br>' +
                f'<b>{nowcast_display_name}</b>: %{{y:.2f}}<extra></extra>'
            ))

        # 添加实际值数据点
        if target_display_name in comparison_df.columns and comparison_df[target_display_name].notna().any():
            actual_plot_data = comparison_df[target_display_name].dropna()
            if not actual_plot_data.empty:
                fig.add_trace(go.Scatter(
                    x=actual_plot_data.index,
                    y=actual_plot_data.values,
                    mode='markers',
                    name=target_display_name,
                    marker=dict(color='red', size=7),
                    hovertemplate=
                    f'<b>日期</b>: %{{x|%Y/%m/%d}}<br>' +
                    f'<b>{target_display_name}</b>: %{{y:.2f}}<extra></extra>'
                ))

        # 添加验证期黄色背景标记
        try:
            # 从metadata获取验证期日期
            validation_start = metadata.get('validation_start_date')
            validation_end = metadata.get('validation_end_date')

            if validation_start and validation_end and validation_start != 'N/A' and validation_end != 'N/A':
                # 转换为datetime对象
                val_start_dt = pd.to_datetime(validation_start)
                val_end_dt = pd.to_datetime(validation_end)

                # 添加黄色半透明背景区域
                fig.add_vrect(
                    x0=val_start_dt,
                    x1=val_end_dt,
                    fillcolor="yellow",
                    opacity=0.2,
                    layer="below",
                    line_width=0,
                    annotation_text="验证期",
                    annotation_position="top left",
                    annotation_font_size=10,
                    annotation_font_color="gray"
                )
                logger.info(f"已添加验证期标记: {val_start_dt} 到 {val_end_dt}")
        except Exception as e:
            logger.warning(f"添加验证期标记失败: {e}")

        # 设置图表布局
        fig.update_layout(
            title=f'周度 {nowcast_display_name} vs. {target_display_name}',
            xaxis=dict(
                title="",
                type='date'
            ),
            yaxis_title="(%)",
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=-0.2,
                xanchor="center",
                x=0.5
            ),
            hovermode='x unified',
            height=500,
            margin=dict(t=50, b=100, l=50, r=50)
        )

        # 显示图表
        st.plotly_chart(fig, use_container_width=True)

        # 提供数据下载
        try:
            csv_data = comparison_df.to_csv(index=True).encode('utf-8-sig')
            st.download_button(
                label="数据下载",
                data=csv_data,
                file_name=f"nowcast_vs_{target_variable_name_for_plot}_aligned.csv",
                mime="text/csv",
                key="download_nowcast_comparison"
            )
        except Exception as e:
            st.error(f"生成下载文件时出错: {e}")

    else:
        st.error("[ERROR] 无法显示Nowcast对比图：未找到complete_aligned_table数据。")

        st.markdown("### [CONFIG] 数据修复选项")

        # 检查是否有原始数据可以重新生成
        has_nowcast = 'calculated_nowcast_orig' in metadata and metadata['calculated_nowcast_orig'] is not None
        has_target = 'original_target_series' in metadata and metadata['original_target_series'] is not None
        has_all_data = 'all_data_aligned_weekly' in metadata and metadata['all_data_aligned_weekly'] is not None
        target_var = metadata.get('target_variable')

        # 显示数据可用性状态
        col1, col2 = st.columns(2)
        with col1:
            st.write("**数据可用性检查:**")
            st.write(f"- calculated_nowcast_orig: {'[SUCCESS]' if has_nowcast else '[ERROR]'}")
            st.write(f"- original_target_series: {'[SUCCESS]' if has_target else '[ERROR]'}")
            st.write(f"- all_data_aligned_weekly: {'[SUCCESS]' if has_all_data else '[ERROR]'}")
            st.write(f"- target_variable: {'[SUCCESS]' if target_var else '[ERROR]'}")

        with col2:
            st.write("**修复建议:**")
            if has_nowcast and has_target:
                st.success("[SUCCESS] 可以从现有数据重新生成complete_aligned_table")
                if st.button("[CONFIG] 立即修复数据", key="repair_data_btn"):
                    try:
                        # 尝试从现有数据重新生成
                        nowcast_data = metadata['calculated_nowcast_orig']
                        target_data = metadata['original_target_series']

                        # 创建对齐表格
                        repaired_df = pd.DataFrame({
                            'Nowcast (Original Scale)': nowcast_data,
                            target_var: target_data
                        })

                        # 只保留有数据的行
                        repaired_df = repaired_df.dropna(how='all')

                        if not repaired_df.empty:
                            # 更新session state中的数据
                            metadata['complete_aligned_table'] = repaired_df
                            st.success(f"[SUCCESS] 数据修复成功！生成了包含 {len(repaired_df)} 行数据的对齐表格")
                            st.info("[LOADING] 请刷新页面查看修复后的图表")

                            # 显示修复后的数据预览
                            st.write("**修复后的数据预览:**")
                            st.dataframe(repaired_df.head(10))
                        else:
                            st.error("[ERROR] 修复失败：生成的数据为空")
                    except Exception as e:
                        st.error(f"[ERROR] 修复过程出错: {e}")
            elif has_all_data and target_var:
                st.warning("[WARNING] 可以从原始数据生成基本对齐表格")
                if st.button("[CONFIG] 生成基本数据", key="generate_basic_btn"):
                    try:
                        all_data = metadata['all_data_aligned_weekly']
                        if target_var in all_data.columns:
                            target_data = all_data[target_var].dropna()
                            if len(target_data) > 0:
                                # 创建基本对齐表格
                                basic_df = pd.DataFrame({
                                    'Nowcast (Original Scale)': target_data,
                                    target_var: target_data
                                })
                                metadata['complete_aligned_table'] = basic_df
                                st.success(f"[SUCCESS] 生成基本数据成功！包含 {len(basic_df)} 行数据")
                                st.info("[LOADING] 请刷新页面查看生成的图表")
                            else:
                                st.error("[ERROR] 目标变量数据为空")
                        else:
                            st.error(f"[ERROR] 目标变量 {target_var} 不在数据中")
                    except Exception as e:
                        st.error(f"[ERROR] 生成过程出错: {e}")
            else:
                st.error("[ERROR] 缺少必要数据，无法修复")
                st.info("[INFO] 建议重新训练模型以生成完整数据")


        comparison_df = None

    # 原 with st.expander("详细分析结果", expanded=False):
    st.markdown("**PCA结果分析**")
    pca_results = metadata.get('pca_results_df')
    # 修正因子数量获取，与训练模块的元数据键匹配
    k = metadata.get('best_params', {}).get('k_factors', metadata.get('n_factors', 0))
    # 兼容numpy整数类型
    if isinstance(k, (int, np.integer)):
        k = int(k)
    if not isinstance(k, int) or k <= 0:
        logger.warning("无法确定最终因子数 k，将尝试从 PCA 数据推断。")
        k = len(pca_results.index) if pca_results is not None and isinstance(pca_results, pd.DataFrame) else 0
    
    if pca_results is not None and isinstance(pca_results, pd.DataFrame) and not pca_results.empty:
        pca_df_display = pca_results.head(k if k > 0 else len(pca_results.index)).copy()
        if '主成分 (PC)' in pca_df_display.columns:
            pca_df_display = pca_df_display.drop(columns=['主成分 (PC)'])
        pca_df_display.insert(0, '主成分 (PC)', [f"PC{i+1}" for i in range(len(pca_df_display.index))])
        if not isinstance(pca_df_display.index, pd.RangeIndex):
            pca_df_display = pca_df_display.reset_index()
            if 'index' in pca_df_display.columns:
                pca_df_display = pca_df_display.rename(columns={'index': 'Original Index'})
        pca_df_display = pca_df_display.rename(columns={
            '解释方差 (%)': '解释方差(%)',
            '累计解释方差 (%)': '累计解释方差(%)',
            '特征值 (Eigenvalue)': '特征值(Eigenvalue)'
        })
        st.dataframe(pca_df_display, use_container_width=True)
    else:
        st.write("未找到 PCA 结果。")
    
    st.markdown("--- ")
    st.markdown("**R² 分析**")
    
    # 创建两列布局显示R²分析表格
    r2_col1, r2_col2 = st.columns(2)
    
    with r2_col1:
        st.markdown("**整体 R² (按行业)**")
        industry_r2 = metadata.get('industry_r2_results')
        if industry_r2 is not None and isinstance(industry_r2, pd.Series) and not industry_r2.empty:
            st.dataframe(industry_r2.to_frame(name="Industry R2 (All Factors)"), use_container_width=True)
            st.caption("附注：衡量所有因子共同解释该行业内所有变量整体变动的百分比。计算方式为对行业内各变量分别对所有因子进行OLS回归后，汇总各变量的总平方和(TSS)与残差平方和(RSS)，计算 Pooled R² = 1 - (Sum(RSS) / Sum(TSS))。")
        else:
            st.write("未找到行业整体 R² 数据。")
         
    with r2_col2:
        st.markdown("**因子对行业 Pooled R²**")
        factor_industry_r2 = metadata.get('factor_industry_r2_results')
        if factor_industry_r2 and isinstance(factor_industry_r2, dict):
            try:
                factor_industry_df = pd.DataFrame(factor_industry_r2)
                st.dataframe(factor_industry_df, use_container_width=True)
                st.caption("附注：衡量单个因子解释该行业内所有变量整体变动的百分比。计算方式为对行业内各变量分别对单个因子进行OLS回归后，汇总TSS与RSS，计算 Pooled R² = 1 - (Sum(RSS) / Sum(TSS))。")
            except ValueError as ve:
                st.warning(f"因子对行业 Pooled R² 数据格式错误，无法转换为DataFrame: {ve}")
                logger.warning(f"Error converting factor_industry_r2 to DataFrame: {factor_industry_r2}")
            except Exception as e:
                st.error(f"显示因子对行业 Pooled R² 时发生未知错误: {e}")
        elif factor_industry_r2 is not None: # It's not a dict or it's an empty dict but not None
            st.write("因子对行业 Pooled R² 数据格式不正确或为空。")
        else:
            st.write("未找到因子对行业 Pooled R² 数据。")

    # 添加数据下载按钮
    # 检查是否有可下载的数据
    has_industry_r2 = industry_r2 is not None and isinstance(industry_r2, pd.Series) and not industry_r2.empty
    has_factor_industry_r2 = (factor_industry_r2 and isinstance(factor_industry_r2, dict) and 
                             len(factor_industry_r2) > 0)
    
    if has_industry_r2 or has_factor_industry_r2:
        try:
            excel_data = generate_r2_excel(industry_r2, factor_industry_r2)
            if excel_data:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"R2_Analysis_Data_{timestamp}.xlsx"
                
                st.download_button(
                    label="数据下载",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="r2_analysis_download_file"
                )
            else:
                st.info("生成Excel文件失败")
        except Exception as e:
            st.error(f"生成下载文件时出错: {str(e)}")
            logger.error(f"Error generating R2 analysis Excel file: {e}")

    st.markdown("---") # Add a separator

    factor_loadings_df = metadata.get('factor_loadings_df') # Assuming this key exists

    if factor_loadings_df is not None and isinstance(factor_loadings_df, pd.DataFrame) and not factor_loadings_df.empty:

        # 使用Backend的聚类函数
        data_for_clustering, y_labels_heatmap, clustering_performed_successfully = perform_loadings_clustering(
            factor_loadings_df,
            cluster_vars=True
        )

        if clustering_performed_successfully:
            logger.info("因子载荷热力图：变量聚类成功。")
        else:
            logger.info("因子载荷热力图：使用原始顺序。")

        factor_names_original = data_for_clustering.columns.tolist()

        # 2. 准备绘图数据 (z值, x轴标签, y轴标签)
        z_values = data_for_clustering.values # (num_clustered_vars, num_factors)
        x_labels_heatmap = factor_names_original # 因子名作为X轴

        fig_heatmap = go.Figure(data=go.Heatmap(
            z=z_values,
            x=x_labels_heatmap,
            y=y_labels_heatmap, # <--- 确保这里使用的是聚类后的 y_labels_heatmap
            colorscale='RdBu_r',  # 反转颜色方案：红色=正相关(正值)，蓝色=负相关(负值)
            zmid=0,
            colorbar=dict(title='载荷值'),
            xgap=1,
            ygap=1,
            hovertemplate=(
                "变量 (Variable): %{y}<br>" +
                "因子 (Factor): %{x}<br>" +
                "载荷值 (Loading): %{z:.4f}<extra></extra>"
            )
        ))

        # Annotate heatmap cells
        annotations = []
        for i, var_name in enumerate(y_labels_heatmap): # 遍历变量 (Y轴)
            for j, factor_name in enumerate(x_labels_heatmap): # 遍历因子 (X轴)
                val = z_values[i][j]
                annotations.append(
                    go.layout.Annotation(
                        text=f"{val:.2f}",
                        x=factor_name,  # X轴是因子
                        y=var_name,     # Y轴是变量
                        xref='x1',
                        yref='y1',
                        showarrow=False,
                        font=dict(color='white' if abs(val) > 0.5 else 'black')
                    )
                )
            
        fig_heatmap.update_layout(
            title="因子载荷聚类热力图 (Factor Loadings Clustermap)", # 修改标题，明确表示聚类功能
            xaxis_title="因子 (Factors)",
            yaxis_title="变量 (Predictors)",
            yaxis=dict( # 确保Y轴按聚类顺序显示
                type='category',
                categoryorder='array', # 强制使用 categoryarray 的顺序
                categoryarray=y_labels_heatmap # <--- 再次确认这里使用的是聚类后的 y_labels_heatmap
            ),
            height=max(600, len(y_labels_heatmap) * 35 + 200),  # 增加高度
            # --- 修改宽度、边距，并将X轴移到顶部 -- -
            width=max(1000, len(x_labels_heatmap) * 100 + max(200, max(len(name) for name in y_labels_heatmap)*8 if y_labels_heatmap else 200) + 50),  # 增加宽度
            margin=dict(l=max(200, max(len(name) for name in y_labels_heatmap)*8 if y_labels_heatmap else 200), r=50, t=100, b=200),  # 增加边距
            annotations=annotations,
            xaxis=dict(
                side='top',       # 将X轴移到顶部
                tickangle=-45    # 保持X轴标签旋转角度
            )
        )
            
        # 使用居中显示的容器
        heatmap_col1, heatmap_col2, heatmap_col3 = st.columns([1, 8, 1])
        with heatmap_col2:
            st.plotly_chart(fig_heatmap, use_container_width=True)

        # Download button for factor loadings data
        try:
            csv_loadings = factor_loadings_df.to_csv(index=True).encode('utf-8-sig') # utf-8-sig for Excel compatibility
            st.download_button(
                label="数据下载",
                data=csv_loadings,
                file_name="factor_loadings.csv",
                mime="text/csv",
                key="download_factor_loadings"
            )
        except Exception as e_csv_loadings:
            st.error(f"生成因子载荷下载文件时出错: {e_csv_loadings}")
        
        st.markdown("---")
        
        # 获取因子时间序列数据
        factor_series_data = metadata.get('factor_series')
        
        if factor_series_data is not None and isinstance(factor_series_data, pd.DataFrame) and not factor_series_data.empty:
            factor_names = factor_series_data.columns.tolist()
            num_factors = len(factor_names)
            
            if num_factors > 0:
                # 确定每行显示的图表数量
                if CONFIG_AVAILABLE:
                    cols_per_row = VisualizationDefaults.FACTOR_PLOT_COLS_EVEN if num_factors % 2 == 0 else VisualizationDefaults.FACTOR_PLOT_COLS_ODD
                else:
                    cols_per_row = 2 if num_factors % 2 == 0 else 3
                
                # 计算需要的行数
                num_rows = (num_factors + cols_per_row - 1) // cols_per_row
                
                # 为每个因子创建时间序列图
                for row in range(num_rows):
                    # 创建列布局
                    cols = st.columns(cols_per_row)
                    
                    for col_idx in range(cols_per_row):
                        factor_idx = row * cols_per_row + col_idx
                        
                        if factor_idx < num_factors:
                            factor_name = factor_names[factor_idx]
                            
                            with cols[col_idx]:
                                # 创建单个因子的时间序列图
                                factor_data = factor_series_data[factor_name].dropna()
                                
                                if not factor_data.empty:
                                    fig_factor = go.Figure()
                                    
                                    fig_factor.add_trace(go.Scatter(
                                        x=factor_data.index,
                                        y=factor_data.values,
                                        mode='lines+markers',
                                        name=factor_name,
                                        line=dict(width=2),
                                        marker=dict(size=4),
                                        hovertemplate=(
                                            f"日期: %{{x|%Y/%m/%d}}<br>" +
                                            f"{factor_name}: %{{y:.4f}}<extra></extra>"
                                        )
                                    ))
                                    
                                    fig_factor.update_layout(
                                        title=f"{factor_name}",
                                        xaxis_title="日期",
                                        yaxis_title="因子值",
                                        height=400,
                                        margin=dict(t=60, b=80, l=60, r=30),
                                        showlegend=False,  # 隐藏图例以节省空间
                                        hovermode='x unified'
                                    )
                                    
                                    # 添加零轴线
                                    fig_factor.add_hline(y=0, line_dash="dash", line_color="gray", opacity=0.5)
                                    
                                    st.plotly_chart(fig_factor, use_container_width=True)
                                else:
                                    st.warning(f"{factor_name}数据为空，无法绘制图表。")
                
                # 提供所有因子数据的统一下载
                
                try:
                    all_factors_csv = factor_series_data.to_csv(index=True).encode('utf-8-sig')
                    st.download_button(
                        label="数据下载",
                        data=all_factors_csv,
                        file_name="所有因子时间序列.csv",
                        mime="text/csv",
                        key="download_all_factors_timeseries"
                    )
                except Exception as e_all_factors:
                    st.error(f"生成所有因子下载文件时出错: {e_all_factors}")
            else:
                st.write("未找到有效的因子数据。")
        else:
            st.write("未在元数据中找到因子时间序列数据。预期的键名: 'factor_series'。")
    
    elif factor_loadings_df is not None and not isinstance(factor_loadings_df, pd.DataFrame):
        st.warning("因子载荷数据 (factor_loadings_df) 存在但不是有效的 DataFrame 格式。")
    else:
        st.write("未在元数据中找到因子载荷数据 (expected key: 'factor_loadings_df')。")


def render_dfm_model_analysis_page(st_module: Any) -> Dict[str, Any]:
    """
    渲染DFM模型分析页面

    Args:
        st_module: Streamlit模块

    Returns:
        Dict[str, Any]: 渲染结果
    """
    try:
        # 调用主要的UI渲染函数
        render_dfm_tab(st_module)

        return {
            'status': 'success',
            'page': 'model_analysis',
            'components': ['file_upload', 'model_info', 'nowcasting', 'factor_analysis']
        }

    except Exception as e:
        st_module.error(f"模型分析页面渲染失败: {str(e)}")
        return {
            'status': 'error',
            'page': 'model_analysis',
            'error': str(e)
        }


def render_dfm_analysis_tab(st_module: Any) -> Dict[str, Any]:
    """
    兼容性接口：渲染DFM模型分析标签页

    Args:
        st_module: Streamlit模块

    Returns:
        Dict[str, Any]: 渲染结果
    """
    return render_dfm_model_analysis_page(st_module)
