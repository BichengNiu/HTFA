# -*- coding: utf-8 -*-
"""
训练结果导出器（合并版）

负责将TrainingResult导出为文件：
- 模型文件（.joblib）
- 元数据文件（.pkl）
- Excel报告（.xlsx）

整合了原metadata_builder、report_generator和utils的功能
"""

import os
import tempfile
import pickle
from datetime import datetime
from typing import Dict, Optional, Any
import numpy as np
import pandas as pd
import joblib
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from dashboard.models.DFM.train.utils.logger import get_logger

logger = get_logger(__name__)


class TrainingResultExporter:
    """训练结果文件导出器（合并版）"""

    def export_all(
        self,
        result,  # TrainingResult
        config,  # TrainingConfig
        output_dir: Optional[str] = None
    ) -> Dict[str, str]:
        """
        导出所有结果文件

        Args:
            result: 训练结果
            config: 训练配置
            output_dir: 输出目录（None=创建临时目录）

        Returns:
            文件路径字典 {
                'final_model_joblib': 模型文件路径,
                'metadata': 元数据文件路径,
                'excel_report': Excel报告路径
            }
        """
        logger.info("开始导出训练结果文件")

        # 创建输出目录
        if output_dir is None:
            output_dir = tempfile.mkdtemp(prefix='dfm_results_')
            logger.info(f"使用临时目录: {output_dir}")
        else:
            os.makedirs(output_dir, exist_ok=True)
            logger.info(f"使用指定目录: {output_dir}")

        # 生成时间戳
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # 导出各个文件
        file_paths = {}

        try:
            model_path = os.path.join(output_dir, f'final_dfm_model_{timestamp}.joblib')
            self._export_model(result, model_path)
            file_paths['final_model_joblib'] = model_path
            logger.info(f"模型文件已导出: {os.path.basename(model_path)}")
        except Exception as e:
            logger.error(f"导出模型文件失败: {e}", exc_info=True)
            file_paths['final_model_joblib'] = None

        try:
            metadata_path = os.path.join(output_dir, f'final_dfm_metadata_{timestamp}.pkl')
            self._export_metadata(result, config, metadata_path, timestamp)
            file_paths['metadata'] = metadata_path
            logger.info(f"元数据文件已导出: {os.path.basename(metadata_path)}")
        except Exception as e:
            logger.error(f"导出元数据文件失败: {e}", exc_info=True)
            file_paths['metadata'] = None

        try:
            excel_path = os.path.join(output_dir, f'final_report_{timestamp}.xlsx')
            self._export_excel_report(result, config, excel_path)
            file_paths['excel_report'] = excel_path
            logger.info(f"Excel报告已导出: {os.path.basename(excel_path)}")
        except Exception as e:
            logger.error(f"导出Excel报告失败: {e}", exc_info=True)
            file_paths['excel_report'] = None

        # 验证文件
        for file_type, path in file_paths.items():
            if path and os.path.exists(path):
                size = os.path.getsize(path)
                logger.debug(f"{file_type}: {path} ({size} bytes)")
            else:
                logger.warning(f"{file_type}: 文件不存在或导出失败")

        logger.info(f"文件导出完成,共 {len([p for p in file_paths.values() if p])} 个文件")
        return file_paths

    def _export_model(self, result, path: str) -> None:
        """导出模型文件"""
        if result.model_result is None:
            raise ValueError("训练结果中没有模型对象")

        joblib.dump(result.model_result, path, compress=3)

        if not os.path.exists(path):
            raise IOError(f"模型文件保存失败: {path}")

        file_size = os.path.getsize(path) / (1024 * 1024)
        logger.debug(f"模型文件大小: {file_size:.2f} MB")

    def _export_metadata(self, result, config, path: str, timestamp: str) -> None:
        """导出元数据文件"""
        metadata = self._build_metadata(result, config, timestamp)
        self._validate_metadata(metadata)

        with open(path, 'wb') as f:
            pickle.dump(metadata, f, protocol=pickle.HIGHEST_PROTOCOL)

        if not os.path.exists(path):
            raise IOError(f"元数据文件保存失败: {path}")

        file_size = os.path.getsize(path) / (1024 * 1024)
        logger.debug(f"元数据文件大小: {file_size:.2f} MB")

    def _export_excel_report(self, result, config, path: str) -> None:
        """导出Excel报告"""
        wb = Workbook()

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # 创建各个sheet
        self._create_params_sheet(wb, result, config)
        self._create_metrics_sheet(wb, result)
        self._create_loadings_sheet(wb, result)
        self._create_transition_matrix_sheet(wb, result)

        wb.save(path)

        if not os.path.exists(path):
            raise IOError(f"Excel报告保存失败: {path}")

        file_size = os.path.getsize(path) / (1024 * 1024)
        logger.debug(f"Excel报告大小: {file_size:.2f} MB")

    # ========== 元数据构建方法 ==========

    def _build_metadata(self, result, config, timestamp: str) -> Dict[str, Any]:
        """构建完整的元数据字典"""
        logger.debug("开始构建元数据")

        metadata = {
            'timestamp': timestamp,
            'target_variable': config.target_variable,
            'best_variables': result.selected_variables,
            'best_params': {
                'k': result.k_factors,
                'max_em_iter': config.max_iterations,
                'tolerance': config.tolerance,
            },
            'train_end_date': getattr(config, 'train_end', ''),
            'validation_start_date': self._calculate_validation_start(config),
            'validation_end_date': getattr(config, 'validation_end', ''),
        }

        # 标准化参数
        if result.model_result and hasattr(result.model_result, 'target_mean'):
            metadata['target_mean_original'] = float(result.model_result.target_mean)
            metadata['target_std_original'] = float(result.model_result.target_std)
        else:
            metadata['target_mean_original'] = 0.0
            metadata['target_std_original'] = 1.0

        # 评估指标
        if result.metrics:
            metadata.update({
                'is_rmse': float(result.metrics.is_rmse),
                'oos_rmse': float(result.metrics.oos_rmse),
                'is_hit_rate': float(result.metrics.is_hit_rate),
                'oos_hit_rate': float(result.metrics.oos_hit_rate),
                'is_correlation': float(result.metrics.is_correlation),
                'oos_correlation': float(result.metrics.oos_correlation),
                # 兼容train_model格式
                'revised_is_rmse': float(result.metrics.is_rmse),
                'revised_oos_rmse': float(result.metrics.oos_rmse),
                'revised_is_hr': float(result.metrics.is_hit_rate),
                'revised_oos_hr': float(result.metrics.oos_hit_rate),
            })

        # 训练统计
        metadata['total_runtime_seconds'] = float(getattr(result, 'training_time', 0.0))
        metadata['factor_loadings_df'] = self._extract_factor_loadings(result)
        metadata['var_industry_map'] = {var: '综合' for var in result.selected_variables}

        # 因子序列
        if result.model_result and hasattr(result.model_result, 'factors'):
            metadata['factor_series'] = result.model_result.factors.copy()

        # 其他字段
        metadata.update({
            'status': 'Success',
            'final_variable_count': len(result.selected_variables),
            'k_factors_final': result.k_factors,
        })

        logger.info(f"元数据构建完成,包含 {len(metadata)} 个字段")
        return metadata

    def _calculate_validation_start(self, config) -> str:
        """计算验证开始日期"""
        try:
            if hasattr(config, 'validation_start') and config.validation_start:
                return config.validation_start

            train_end_value = getattr(config, 'train_end', None)
            if train_end_value:
                train_end = pd.to_datetime(train_end_value)
                validation_start = train_end + pd.DateOffset(weeks=1)
                return validation_start.strftime('%Y-%m-%d')

            return ''
        except Exception as e:
            logger.warning(f"计算验证开始日期失败: {e}")
            return ''

    def _validate_metadata(self, metadata: Dict) -> None:
        """验证元数据包含所有必需字段"""
        required_fields = [
            'timestamp', 'target_variable', 'best_variables',
            'best_params', 'train_end_date', 'validation_end_date',
        ]

        missing_fields = [f for f in required_fields if f not in metadata]

        if missing_fields:
            raise ValueError(f"元数据缺少必需字段: {missing_fields}")

        logger.debug(f"元数据验证通过,包含 {len(metadata)} 个字段")

    # ========== Excel报告生成方法 ==========

    def _create_params_sheet(self, wb: Workbook, result, config) -> None:
        """创建模型参数sheet"""
        ws = wb.create_sheet("模型参数")

        train_end = getattr(config, 'train_end', 'N/A')
        validation_end = getattr(config, 'validation_end', 'N/A')

        params_data = [
            ['参数名', '参数值'],
            ['因子数', result.k_factors],
            ['EM最大迭代次数', config.max_iterations],
            ['收敛容差', config.tolerance],
            ['训练结束日期', train_end],
            ['验证结束日期', validation_end],
            ['目标变量', config.target_variable],
            ['选中变量数', len(result.selected_variables)],
            ['', ''],
            ['选中变量列表', ''],
        ]

        for i, var in enumerate(result.selected_variables, 1):
            params_data.append([f'变量{i}', var])

        for row in params_data:
            ws.append(row)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

    def _create_metrics_sheet(self, wb: Workbook, result) -> None:
        """创建评估指标sheet"""
        ws = wb.create_sheet("评估指标")

        if result.metrics is None:
            ws.append(['评估指标', '值'])
            return

        metrics_data = [
            ['评估指标', '样本内(IS)', '样本外(OOS)'],
            ['RMSE', f'{result.metrics.is_rmse:.4f}', f'{result.metrics.oos_rmse:.4f}'],
            ['Hit Rate', f'{result.metrics.is_hit_rate:.2%}', f'{result.metrics.oos_hit_rate:.2%}'],
            ['相关系数', f'{result.metrics.is_correlation:.4f}', f'{result.metrics.oos_correlation:.4f}'],
        ]

        if hasattr(result, 'training_time'):
            metrics_data.append(['', '', ''])
            metrics_data.append(['训练耗时(秒)', f'{result.training_time:.2f}', ''])

        for row in metrics_data:
            ws.append(row)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def _create_loadings_sheet(self, wb: Workbook, result) -> None:
        """创建因子载荷sheet"""
        ws = wb.create_sheet("因子载荷")

        loadings_df = self._extract_factor_loadings(result)

        if loadings_df.empty:
            ws.append(['变量', '载荷'])
            return

        for r in dataframe_to_rows(loadings_df, index=True, header=True):
            ws.append(r)

        ws.column_dimensions['A'].width = 30
        for col in range(2, result.k_factors + 2):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _create_transition_matrix_sheet(self, wb: Workbook, result) -> None:
        """创建状态转移矩阵sheet"""
        if not result.model_result:
            return

        A = result.model_result.A

        if A is None:
            return

        ws = wb.create_sheet("状态转移矩阵")

        if isinstance(A, np.ndarray):
            factor_names = [f'Factor_{i+1}' for i in range(A.shape[0])]
            A_df = pd.DataFrame(A, index=factor_names, columns=factor_names)
        elif isinstance(A, pd.DataFrame):
            A_df = A
        else:
            return

        for r in dataframe_to_rows(A_df, index=True, header=True):
            ws.append(r)

        for col in range(1, A_df.shape[1] + 2):
            ws.column_dimensions[chr(64 + col)].width = 15

    # ========== 工具方法 ==========

    def _extract_factor_loadings(self, result) -> pd.DataFrame:
        """提取因子载荷矩阵（H矩阵）"""
        try:
            if not result.model_result:
                return pd.DataFrame()

            H = result.model_result.H

            if H is None:
                return pd.DataFrame()

            if isinstance(H, np.ndarray):
                factor_names = [f'Factor_{i+1}' for i in range(H.shape[1])]
                return pd.DataFrame(H, columns=factor_names, index=result.selected_variables)
            elif isinstance(H, pd.DataFrame):
                return H.copy()

            return pd.DataFrame()

        except Exception as e:
            logger.error(f"提取因子载荷失败: {e}")
            return pd.DataFrame()


__all__ = ['TrainingResultExporter']
