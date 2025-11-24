"""
Excel Generator Utilities
Excel生成工具
"""

import io
import logging
from datetime import datetime
from typing import Optional, Dict
import pandas as pd
from openpyxl import Workbook

logger = logging.getLogger(__name__)


class R2ExcelGenerator:
    """
    R²分析数据Excel生成器

    封装generate_r2_excel函数的逻辑，提供更好的可维护性
    """

    def __init__(self, industry_r2: pd.Series, factor_industry_r2: Dict):
        """
        初始化生成器

        Args:
            industry_r2: 整体R²数据(按行业)
            factor_industry_r2: 因子对行业Pooled R²数据
        """
        self.industry_r2 = industry_r2
        self.factor_industry_r2 = factor_industry_r2
        self.logger = logging.getLogger(self.__class__.__name__)

    def generate(self) -> Optional[bytes]:
        """
        生成Excel文件

        Returns:
            Excel文件的字节数据，失败时返回None
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "R2_Analysis"

            current_row = 1

            # 添加标题
            self._add_title(ws, current_row)
            current_row += 2

            # 添加生成时间
            ws.cell(row=current_row, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            current_row += 3

            # 第一个表：整体R²(按行业)
            current_row = self._add_industry_r2_table(ws, current_row)

            # 第二个表：因子对行业Pooled R²
            current_row = self._add_factor_industry_r2_table(ws, current_row)

            # 调整列宽
            self._adjust_column_widths(ws)

            # 保存到字节流
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except Exception as e:
            self.logger.error(f"生成R²Excel文件时发生错误: {e}")
            return None

    def _add_title(self, ws, row: int) -> None:
        """添加标题"""
        ws.cell(row=row, column=1, value="R² 分析数据报告")
        ws.cell(row=row, column=1).font = ws.cell(row=row, column=1).font.copy(bold=True, size=14)

    def _add_industry_r2_table(self, ws, start_row: int) -> int:
        """
        添加整体R²表

        Returns:
            下一个可用行号
        """
        if self.industry_r2 is None or not isinstance(self.industry_r2, pd.Series) or self.industry_r2.empty:
            return start_row

        current_row = start_row

        # 表标题
        ws.cell(row=current_row, column=1, value="整体 R² (按行业)")
        ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=12)
        current_row += 2

        # 表头
        ws.cell(row=current_row, column=1, value="行业")
        ws.cell(row=current_row, column=2, value="Industry R2 (All Factors)")
        for col in range(1, 3):
            ws.cell(row=current_row, column=col).font = ws.cell(row=current_row, column=col).font.copy(bold=True)
        current_row += 1

        # 数据行
        for index, value in self.industry_r2.items():
            ws.cell(row=current_row, column=1, value=str(index))
            ws.cell(row=current_row, column=2, value=float(value) if pd.notna(value) else None)
            current_row += 1

        # 附注
        current_row = self._add_industry_r2_notes(ws, current_row)

        return current_row

    def _add_industry_r2_notes(self, ws, start_row: int) -> int:
        """添加整体R²附注"""
        current_row = start_row + 1
        ws.cell(row=current_row, column=1, value="附注：")
        ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True)
        current_row += 1

        notes = [
            "衡量所有因子共同解释该行业内所有变量整体变动的百分比。",
            "计算方式为对行业内各变量分别对所有因子进行OLS回归后，",
            "汇总各变量的总平方和(TSS)与残差平方和(RSS)，",
            "计算 Pooled R² = 1 - (Sum(RSS) / Sum(TSS))。"
        ]

        for note in notes:
            ws.cell(row=current_row, column=1, value=note)
            current_row += 1

        return current_row + 2

    def _add_factor_industry_r2_table(self, ws, start_row: int) -> int:
        """
        添加因子对行业Pooled R²表

        Returns:
            下一个可用行号
        """
        if not self.factor_industry_r2 or not isinstance(self.factor_industry_r2, dict) or len(self.factor_industry_r2) == 0:
            return start_row

        try:
            factor_industry_df = pd.DataFrame(self.factor_industry_r2)

            current_row = start_row

            # 表标题
            ws.cell(row=current_row, column=1, value="因子对行业 Pooled R²")
            ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True, size=12)
            current_row += 2

            # 表头
            for col_idx, column in enumerate(['行业/因子'] + list(factor_industry_df.columns)):
                ws.cell(row=current_row, column=col_idx + 1, value=column)
                ws.cell(row=current_row, column=col_idx + 1).font = ws.cell(row=current_row, column=col_idx + 1).font.copy(bold=True)
            current_row += 1

            # 数据行
            for index, row_data in factor_industry_df.iterrows():
                ws.cell(row=current_row, column=1, value=str(index))
                for col_idx, value in enumerate(row_data):
                    ws.cell(row=current_row, column=col_idx + 2, value=float(value) if pd.notna(value) else None)
                current_row += 1

            # 附注
            current_row = self._add_factor_industry_r2_notes(ws, current_row)

            return current_row

        except Exception as e:
            self.logger.error(f"处理factor_industry_r2数据时出错: {e}")
            return start_row

    def _add_factor_industry_r2_notes(self, ws, start_row: int) -> int:
        """添加因子对行业R²附注"""
        current_row = start_row + 1
        ws.cell(row=current_row, column=1, value="附注：")
        ws.cell(row=current_row, column=1).font = ws.cell(row=current_row, column=1).font.copy(bold=True)
        current_row += 1

        notes = [
            "衡量单个因子解释该行业内所有变量整体变动的百分比。",
            "计算方式为对行业内各变量分别对单个因子进行OLS回归后，",
            "汇总TSS与RSS，计算 Pooled R² = 1 - (Sum(RSS) / Sum(TSS))。"
        ]

        for note in notes:
            ws.cell(row=current_row, column=1, value=note)
            current_row += 1

        return current_row + 2

    @staticmethod
    def _adjust_column_widths(ws) -> None:
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def generate_r2_excel(industry_r2: pd.Series, factor_industry_r2: dict) -> Optional[bytes]:
    """
    生成R²分析数据的Excel文件（向后兼容函数）

    Args:
        industry_r2: 整体R²数据(按行业)
        factor_industry_r2: 因子对行业Pooled R²数据

    Returns:
        bytes: Excel文件的字节数据
    """
    generator = R2ExcelGenerator(industry_r2, factor_industry_r2)
    return generator.generate()
